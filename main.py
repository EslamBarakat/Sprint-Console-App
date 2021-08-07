# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable
#First Name -fname
#Second Name -sname
#Email-email
#Phone -Phone
#Password
#Confirmation Password-cpassword

# opening the existing excel file
wb = load_workbook('C:\\Users\\DELL\\Desktop\\excel.xlsx')

# create the sheet object
sheet = wb.active


def excel():

    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Firs Name"
    sheet.cell(row=1, column=2).value = "Second Name "
    sheet.cell(row=1, column=3).value = "Email"
    sheet.cell(row=1, column=4).value = "Phone"
    sheet.cell(row=1, column=5).value = "Password"
    sheet.cell(row=1, column=6).value = "Confirmation Password"


# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    fname_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    sname_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    email_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
   phone_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    password_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    cpassword_field.focus_set()



# Function for clearing the
# contents of text entry boxes
def clear():

    # clear the content of text entry box
    fname_field.delete(0, END)
    sname_field.delete(0, END)
    email_field.delete(0, END)
    phone_field.delete(0, END)
    password_field.delete(0, END)
    cpassword_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file


def insert():

    # if user not fill any entry
    # then print "empty input"
    if (fname_field.get() == "" or
            sname_field.get() == "" or
            email_field.get() == "" or
            phone_field.get() == "" or
            password_field.get() == "" or
            cpassword_field.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = fname_field.get()
        sheet.cell(row=current_row + 1, column=2).value = sname_field.get()
        sheet.cell(row=current_row + 1, column=3).value = email_field.get()
        sheet.cell(row=current_row + 1, column=4).value = phone_field.get()
        sheet.cell(row=current_row + 1, column=5).value = password_field.get()
        sheet.cell(row=current_row + 1, column=6).value = cpassword_field.get()

        # save the file
        wb.save('C:\\Users\\DELL\\Desktop\\excel.xlsx')


        # set focus on the name_field box
        fname_field.focus_set()

        # call the clear() function
        clear()


# Driver code
if __name__ == "__main__":

    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("registration form")

    # set the configuration of GUI window
    root.geometry("500x300")

    excel()

    # create a Form label
    heading = Label(root, text="Form", bg="white")

    # create a Name label
    fname = Label(root, text="First Name", bg="light green")

    # create a Course label
    sname = Label(root, text="Second Name", bg="light green")

    # create a Semester label
    email = Label(root, text="Email", bg="light green")

    # create a Form No. lable
    phone = Label(root, text="Phone", bg="light green")

    # create a Contact No. label
    password = Label(root, text="Password", bg="light green")

    # create a Email id label
    cpassword = Label(root, text="Confirmation Password", bg="light green")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    fname.grid(row=1, column=0)
    sname.grid(row=2, column=0)
    email.grid(row=3, column=0)
    phone.grid(row=4, column=0)
    password.grid(row=5, column=0)
    cpassword.grid(row=6, column=0)

    # create a text entry box
    # for typing the information
    fname_field = Entry(root)
    sname_field = Entry(root)
    email_field = Entry(root)
    phone_field = Entry(root)
    password_field = Entry(root)
    cpassword_field = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    fname_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    sname_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    email_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    phone_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    password_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    cpassword_field.bind("<Return>", focus6)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    fname_field.grid(row=1, column=1, ipadx="100")
    sname_field.grid(row=2, column=1, ipadx="100")
    email_field.grid(row=3, column=1, ipadx="100")
    phone_field.grid(row=4, column=1, ipadx="100")
    password_field.grid(row=5, column=1, ipadx="100")
    cpassword_field.grid(row=6, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=8, column=1)

    # start the GUI
    root.mainloop()